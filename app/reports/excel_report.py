"""Export Excel des rapports et des ventes via openpyxl."""

from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from app import config
from app.services import settings_service

_HEADER_FILL = PatternFill("solid", fgColor="2563EB")
_HEADER_FONT = Font(bold=True, color="FFFFFF")


def _style_header(worksheet, ncols: int) -> None:
    for col in range(1, ncols + 1):
        cell = worksheet.cell(row=1, column=col)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center")


def _autosize(worksheet) -> None:
    for column_cells in worksheet.columns:
        length = max((len(str(c.value or "")) for c in column_cells), default=10)
        letter = column_cells[0].column_letter
        worksheet.column_dimensions[letter].width = min(45, length + 4)


def export_report_excel(
    report: dict, sales_rows=None, path: str | Path | None = None
) -> Path:
    """Génère un classeur Excel : synthèse + détail des ventes."""
    config.ensure_directories()
    shop = settings_service.get_shop_info()
    currency = shop.currency or "FCFA"

    if path is None:
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        path = config.EXPORT_DIR / f"rapport_{stamp}.xlsx"
    path = Path(path)

    workbook = Workbook()

    summary = workbook.active
    summary.title = "Synthèse"
    summary.append(["Indicateur", f"Valeur ({currency})"])
    summary.append(["Période", f"{report['start']:%d/%m/%Y} - {report['end']:%d/%m/%Y}"])
    summary.append(["Chiffre d'affaires", report["revenue"]])
    summary.append(["Nombre de ventes", report["sales_count"]])
    summary.append(["Bénéfice brut", report["profit"]])
    summary.append(["Dépenses", report["expenses"]])
    summary.append(["Bénéfice net", report["net_profit"]])
    _style_header(summary, 2)
    _autosize(summary)

    top = workbook.create_sheet("Top produits")
    top.append(["Produit", "Quantité", "Chiffre d'affaires"])
    for name, qty, total in report["top_products"]:
        top.append([name, qty, total])
    _style_header(top, 3)
    _autosize(top)

    if sales_rows:
        detail = workbook.create_sheet("Ventes")
        detail.append(["Ticket", "Date", "Total", "Bénéfice", "Statut"])
        for row in sales_rows:
            detail.append(list(row))
        _style_header(detail, 5)
        _autosize(detail)

    workbook.save(path)
    return path


def export_products_excel(products, path: str | Path | None = None) -> Path:
    """Exporte la liste des produits (utile pour l'inventaire)."""
    config.ensure_directories()
    if path is None:
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        path = config.EXPORT_DIR / f"produits_{stamp}.xlsx"
    path = Path(path)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Produits"
    sheet.append(
        ["Nom", "Catégorie", "Code-barres", "Référence", "Prix achat", "Prix vente", "Quantité", "Unité"]
    )
    for p in products:
        sheet.append(
            [
                p.name,
                p.category_name,
                p.barcode,
                p.reference,
                float(p.purchase_price),
                float(p.sale_price),
                float(p.quantity),
                p.unit_name,
            ]
        )
    _style_header(sheet, 8)
    _autosize(sheet)
    workbook.save(path)
    return path
